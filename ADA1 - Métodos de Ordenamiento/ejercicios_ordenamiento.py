"""
EJERCICIOS DE ORDENAMIENTO - 3 MÉTODOS EN UN SOLO CÓDIGO

Ejercicio 1: Método Burbuja (Bubble Sort)
Ejercicio 2: Método Inserción (Insertion Sort)
Ejercicio 3: Método Selección (Selection Sort)

Características:
- Puede ordenar datos desde 1 archivo O 2 archivos (combinados)
- Soporta formatos: TXT, JSON, Excel (.xlsx)
- Guarda el resultado ordenado

Autor: Estudiante
Fecha: 2024
"""

import os
import json
from datetime import datetime

# =============================================================
# LIBRERÍAS OPCIONALES
# =============================================================

try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

try:
    import tkinter as tk
    from tkinter import filedialog
    TKINTER_DISPONIBLE = True
except ImportError:
    TKINTER_DISPONIBLE = False


# =============================================================
# MÉTODO 1: BURBUJA (Bubble Sort)
# =============================================================

def burbuja(lista):
    """
    Ordenamiento por Burbuja
    Complejidad: O(n²)
    Compara elementos adyacentes y los intercambia si están en orden incorrecto
    """
    lista = lista[:]
    n = len(lista)
    for i in range(n - 1):
        intercambiado = False
        for j in range(n - 1 - i):
            if lista[j] > lista[j + 1]:
                lista[j], lista[j + 1] = lista[j + 1], lista[j]
                intercambiado = True
        if not intercambiado:
            break
    return lista


# =============================================================
# MÉTODO 2: INSERCIÓN (Insertion Sort)
# =============================================================

def insercion(lista):
    """
    Ordenamiento por Inserción
    Complejidad: O(n²)
    Toma cada elemento y lo inserta en su posición correcta en la parte ordenada
    """
    lista = lista[:]
    for i in range(1, len(lista)):
        clave = lista[i]
        j = i - 1
        while j >= 0 and lista[j] > clave:
            lista[j + 1] = lista[j]
            j -= 1
        lista[j + 1] = clave
    return lista


# =============================================================
# MÉTODO 3: SELECCIÓN (Selection Sort)
# =============================================================

def seleccion(lista):
    """
    Ordenamiento por Selección
    Complejidad: O(n²)
    Encuentra el elemento mínimo y lo coloca al inicio
    """
    lista = lista[:]
    n = len(lista)
    for i in range(n - 1):
        min_idx = i
        for j in range(i + 1, n):
            if lista[j] < lista[min_idx]:
                min_idx = j
        if min_idx != i:
            lista[i], lista[min_idx] = lista[min_idx], lista[i]
    return lista


# =============================================================
# FUNCIONES DE LECTURA DE ARCHIVOS
# =============================================================

def leer_txt(ruta):
    """Lee números de un archivo TXT"""
    with open(ruta, 'r', encoding='utf-8') as f:
        contenido = f.read()
    
    tokens = contenido.replace(',', ' ').replace('\n', ' ').split()
    numeros = []
    
    for token in tokens:
        try:
            num = float(token)
            numeros.append(int(num) if num == int(num) else num)
        except ValueError:
            print(f"    [!] Advertencia: '{token}' no es número, se omite")
    
    if not numeros:
        raise ValueError("No se encontraron números en el archivo")
    
    return numeros


def leer_json(ruta):
    """Lee números de un archivo JSON"""
    with open(ruta, 'r', encoding='utf-8') as f:
        datos = json.load(f)
    
    numeros = []
    
    def extraer(obj):
        if isinstance(obj, (int, float)):
            numeros.append(int(obj) if obj == int(obj) else obj)
        elif isinstance(obj, list):
            for item in obj:
                extraer(item)
        elif isinstance(obj, dict):
            for valor in obj.values():
                extraer(valor)
    
    extraer(datos)
    
    if not numeros:
        raise ValueError("No se encontraron números en el JSON")
    
    return numeros


def leer_excel(ruta):
    """Lee números de un archivo Excel"""
    if not EXCEL_DISPONIBLE:
        raise ImportError("Instala openpyxl: pip install openpyxl")
    
    wb = openpyxl.load_workbook(ruta, data_only=True)
    numeros = []
    
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(values_only=True):
            for celda in fila:
                if isinstance(celda, (int, float)) and celda is not None:
                    num = int(celda) if celda == int(celda) else celda
                    numeros.append(num)
    
    wb.close()
    
    if not numeros:
        raise ValueError("No se encontraron números en el Excel")
    
    return numeros


def leer_archivo(ruta):
    """Detecta el tipo de archivo y lee los números"""
    ext = os.path.splitext(ruta)[1].lower()
    
    if ext == '.txt':
        return leer_txt(ruta)
    elif ext == '.json':
        return leer_json(ruta)
    elif ext in ('.xlsx', '.xls'):
        return leer_excel(ruta)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Usa .txt, .json o .xlsx")


# =============================================================
# FUNCIONES DE SELECCIÓN DE ARCHIVOS
# =============================================================

def seleccionar_archivo(mensaje, obligatorio=True):
    """Selecciona archivo usando explorador o ruta manual"""
    if TKINTER_DISPONIBLE:
        print(f"\n  {mensaje}")
        if not obligatorio:
            print("    (Puedes dejar vacío si solo quieres 1 archivo)")
        print("    E = Usar explorador de archivos")
        print("    M = Escribir ruta manualmente")
        print("    N = No usar más archivos (solo 1 archivo)")
        opcion = input("  Elige (E/M/N): ").strip().upper()
        
        if opcion == 'N' and not obligatorio:
            return None
        
        if opcion == 'E':
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            ruta = filedialog.askopenfilename(
                title=mensaje,
                filetypes=[
                    ("Archivos soportados", "*.txt *.json *.xlsx *.xls"),
                    ("Texto", "*.txt"),
                    ("JSON", "*.json"),
                    ("Excel", "*.xlsx *.xls"),
                ]
            )
            root.destroy()
            if ruta:
                return ruta
            print("  [!] No seleccionaste archivo, usa modo manual")
    
    # Modo manual
    while True:
        ruta = input(f"  {mensaje} (dejar vacío si no quieres más archivos): ").strip().strip('"')
        if not ruta and not obligatorio:
            return None
        if not ruta:
            continue
        if os.path.isfile(ruta):
            return ruta
        print(f"  [!] Archivo no encontrado: {ruta}")


def guardar_resultado(numeros, metodo, archivos):
    """Guarda el resultado en un archivo"""
    nombre_base = f"resultado_{metodo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    
    if TKINTER_DISPONIBLE:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        ruta = filedialog.asksaveasfilename(
            title="Guardar resultado como...",
            defaultextension=".txt",
            initialfile=nombre_base,
            filetypes=[("Texto", "*.txt")]
        )
        root.destroy()
        if not ruta:
            ruta = nombre_base
    else:
        ruta = input(f"  Nombre del archivo de salida [default: {nombre_base}]: ").strip()
        if not ruta:
            ruta = nombre_base
        if not ruta.endswith('.txt'):
            ruta += '.txt'
    
    with open(ruta, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write(f"  RESULTADO DE ORDENAMIENTO - {metodo.upper()}\n")
        f.write("=" * 60 + "\n")
        f.write(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"  Método: {metodo}\n")
        f.write(f"  Archivos utilizados: {len(archivos)}\n")
        for i, arch in enumerate(archivos, 1):
            f.write(f"    Archivo {i}: {os.path.basename(arch)}\n")
        f.write(f"  Total números: {len(numeros)}\n")
        f.write("=" * 60 + "\n\n")
        f.write("Números originales (combinados):\n")
        f.write("-" * 40 + "\n")
        f.write("Números ordenados:\n")
        f.write("-" * 40 + "\n")
        f.write(", ".join(str(n) for n in numeros) + "\n")
    
    print(f"\n  ✅ Resultado guardado en: {os.path.abspath(ruta)}")
    return ruta


# =============================================================
# FUNCIÓN GENÉRICA PARA EJECUTAR CUALQUIER MÉTODO
# =============================================================

def ejecutar_ejercicio(nombre_metodo, funcion_ordenamiento):
    """
    Ejecuta un ejercicio de ordenamiento con 1 o 2 archivos
    """
    print("\n" + "=" * 60)
    print(f"  EJERCICIO: {nombre_metodo}")
    print("=" * 60)
    
    # Mostrar explicación del método
    explicaciones = {
        "Burbuja": "Compara elementos adyacentes y los intercambia si están en orden incorrecto. El elemento más grande 'burbujea' hacia el final.",
        "Inserción": "Toma cada elemento y lo inserta en su posición correcta dentro de la parte ya ordenada de la lista.",
        "Selección": "Encuentra el elemento mínimo en la parte no ordenada y lo intercambia con el primer elemento no ordenado."
    }
    
    print(f"\n  📖 Explicación del método {nombre_metodo}:")
    print(f"     {explicaciones.get(nombre_metodo, '')}")
    print(f"     Complejidad: O(n²)")
    
    # Seleccionar archivos
    print("\n  📁 SELECCIÓN DE ARCHIVOS:")
    print("  ─────────────────────────")
    
    archivo1 = seleccionar_archivo("Selecciona el PRIMER archivo", obligatorio=True)
    if not archivo1:
        print("  ❌ Debes seleccionar al menos un archivo")
        return
    
    print(f"\n  ✅ Primer archivo: {os.path.basename(archivo1)}")
    
    # Preguntar si quiere un segundo archivo
    print("\n  ¿Deseas agregar un SEGUNDO archivo para combinarlo?")
    respuesta = input("  (s/n): ").strip().lower()
    
    archivo2 = None
    if respuesta == 's':
        archivo2 = seleccionar_archivo("Selecciona el SEGUNDO archivo", obligatorio=False)
        if archivo2:
            print(f"  ✅ Segundo archivo: {os.path.basename(archivo2)}")
    
    # Leer datos
    try:
        datos1 = leer_archivo(archivo1)
        print(f"\n  📁 {os.path.basename(archivo1)}: {len(datos1)} números")
        print(f"     {datos1}")
        
        archivos_usados = [archivo1]
        
        if archivo2:
            datos2 = leer_archivo(archivo2)
            print(f"\n  📁 {os.path.basename(archivo2)}: {len(datos2)} números")
            print(f"     {datos2}")
            combinados = datos1 + datos2
            archivos_usados.append(archivo2)
            print(f"\n  📊 Datos combinados ({len(combinados)} números):")
        else:
            combinados = datos1
            print(f"\n  📊 Datos del archivo ({len(combinados)} números):")
        
        print(f"     {combinados}")
        
        # Ordenar
        print(f"\n  🔄 Ordenando con MÉTODO {nombre_metodo.upper()}...")
        resultado = funcion_ordenamiento(combinados)
        
        print(f"\n  ✅ Resultado ordenado ({len(resultado)} números):")
        print(f"     {resultado}")
        
        # Guardar resultado
        guardar_resultado(resultado, nombre_metodo.lower(), archivos_usados)
        
    except Exception as e:
        print(f"\n  ❌ Error: {e}")


# =============================================================
# CREAR ARCHIVOS DE EJEMPLO
# =============================================================

def crear_archivos_ejemplo():
    """Crea archivos de ejemplo para probar los ejercicios"""
    print("\n  📁 Creando archivos de ejemplo...")
    
    # Archivo 1: datos desordenados
    with open("datos1.txt", "w", encoding="utf-8") as f:
        f.write("85, 72, 91, 60, 78, 34, 55, 19")
    
    with open("datos1.json", "w", encoding="utf-8") as f:
        json.dump([85, 72, 91, 60, 78, 34, 55, 19], f)
    
    # Archivo 2: más datos desordenados
    with open("datos2.txt", "w", encoding="utf-8") as f:
        f.write("43, 67, 11, 98, 25, 50, 73, 88")
    
    with open("datos2.json", "w", encoding="utf-8") as f:
        json.dump({"numeros": [43, 67, 11, 98, 25, 50, 73, 88]}, f)
    
    # Archivo 3: datos adicionales
    with open("datos3.txt", "w", encoding="utf-8") as f:
        f.write("5, 2, 9, 1, 7, 3, 8, 4, 6")
    
    # Excel
    if EXCEL_DISPONIBLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Numeros"
        for v in [85, 72, 91, 60, 78, 34, 55, 19]:
            ws.append([v])
        wb.save("datos1.xlsx")
        
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = "Valores"
        for v in [43, 67, 11, 98, 25, 50, 73, 88]:
            ws2.append([v])
        wb2.save("datos2.xlsx")
    
    print("\n  ✅ Archivos creados:")
    print("     📄 datos1.txt, datos2.txt, datos3.txt")
    print("     📄 datos1.json, datos2.json")
    if EXCEL_DISPONIBLE:
        print("     📊 datos1.xlsx, datos2.xlsx")


# =============================================================
# MENÚ PRINCIPAL
# =============================================================

def mostrar_menu():
    print("=" * 60)
    print("   SISTEMA DE ORDENAMIENTO - 3 EJERCICIOS")
    print("=" * 60)
    print("""
   ┌─────────────────────────────────────────────────────────┐
   │  EJERCICIO 1: MÉTODO BURBUJA                            │
   │  └─ Ordena datos comparando elementos adyacentes        │
   │                                                         │
   │  EJERCICIO 2: MÉTODO INSERCIÓN                          │
   │  └─ Ordena datos insertando cada elemento en su lugar   │
   │                                                         │
   │  EJERCICIO 3: MÉTODO SELECCIÓN                          │
   │  └─ Ordena datos seleccionando el mínimo repetidamente  │
   └─────────────────────────────────────────────────────────┘
   """)
    print("   Opciones adicionales:")
    print("     4. Crear archivos de ejemplo")
    print("     5. Ver explicación de métodos")
    print("     0. Salir")
    print("-" * 60)


def mostrar_explicaciones():
    """Muestra explicación detallada de cada método"""
    print("\n" + "=" * 60)
    print("  EXPLICACIÓN DE MÉTODOS DE ORDENAMIENTO")
    print("=" * 60)
    
    print("""
┌─────────────────────────────────────────────────────────────────┐
│ 1. MÉTODO BURBUJA (Bubble Sort)                                 │
├─────────────────────────────────────────────────────────────────┤
│ • Funcionamiento:                                               │
│   Compara pares de elementos adyacentes y los intercambia       │
│   si están en el orden incorrecto.                              │
│ • Ejemplo: [5, 3, 8, 1] → [3, 5, 1, 8] → [3, 1, 5, 8] → etc    │
│ • Complejidad: O(n²) en el peor caso                            │
│ • Estable: Sí                                                   │
└─────────────────────────────────────────────────────────────────┘

┌─────────────────────────────────────────────────────────────────┐
│ 2. MÉTODO INSERCIÓN (Insertion Sort)                            │
├─────────────────────────────────────────────────────────────────┤
│ • Funcionamiento:                                               │
│   Toma cada elemento y lo inserta en su posición correcta       │
│   dentro de la parte ya ordenada de la lista.                   │
│ • Ejemplo: [5, 3, 8, 1] → [3, 5, 8, 1] → [3, 5, 8, 1] → etc    │
│ • Complejidad: O(n²) en el peor caso                            │
│ • Estable: Sí                                                   │
└─────────────────────────────────────────────────────────────────┘

┌─────────────────────────────────────────────────────────────────┐
│ 3. MÉTODO SELECCIÓN (Selection Sort)                            │
├─────────────────────────────────────────────────────────────────┤
│ • Funcionamiento:                                               │
│   Encuentra el elemento mínimo y lo coloca al inicio, luego     │
│   busca el siguiente mínimo, y así sucesivamente.               │
│ • Ejemplo: [5, 3, 8, 1] → [1, 3, 8, 5] → [1, 3, 5, 8]          │
│ • Complejidad: O(n²) en todos los casos                         │
│ • Estable: No (puede cambiar el orden de elementos iguales)     │
└─────────────────────────────────────────────────────────────────┘

📌 COMPLEJIDADES COMPARADAS:
   • Los 3 métodos tienen complejidad O(n²)
   • Para listas pequeñas (< 100 elementos) son aceptables
   • Para listas grandes, usar métodos avanzados (Quick, Merge, Heap)
""")


def limpiar():
    os.system('cls' if os.name == 'nt' else 'clear')


def main():
    limpiar()
    print("=" * 60)
    print("   BIENVENIDO - EJERCICIOS DE ORDENAMIENTO")
    print("=" * 60)
    print("\n  Este programa contiene 3 ejercicios que demuestran")
    print("  los métodos de ordenamiento: Burbuja, Inserción y Selección.")
    print("\n  ✅ Puedes usar 1 archivo o combinar 2 archivos")
    print("  ✅ Soporta formatos: TXT, JSON, Excel (.xlsx)")
    print("  ✅ Guarda el resultado ordenado\n")
    
    respuesta = input("  ¿Deseas crear archivos de ejemplo? (s/n): ").strip().lower()
    if respuesta == 's':
        crear_archivos_ejemplo()
    
    input("\n  ⏎ Presiona ENTER para continuar...")
    
    while True:
        limpiar()
        mostrar_menu()
        
        opcion = input("\n  Selecciona una opción (0-5): ").strip()
        
        if opcion == '1':
            ejecutar_ejercicio("Burbuja", burbuja)
            input("\n  ⏎ Presiona ENTER para continuar...")
        elif opcion == '2':
            ejecutar_ejercicio("Inserción", insercion)
            input("\n  ⏎ Presiona ENTER para continuar...")
        elif opcion == '3':
            ejecutar_ejercicio("Selección", seleccion)
            input("\n  ⏎ Presiona ENTER para continuar...")
        elif opcion == '4':
            crear_archivos_ejemplo()
            input("\n  ⏎ Presiona ENTER para continuar...")
        elif opcion == '5':
            mostrar_explicaciones()
            input("\n  ⏎ Presiona ENTER para continuar...")
        elif opcion == '0':
            print("\n  👋 ¡Hasta luego!\n")
            break
        else:
            print("\n  ❌ Opción no válida")
            input("\n  ⏎ Presiona ENTER para continuar...")


if __name__ == "__main__":
    main()
