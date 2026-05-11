"""
LIBRERÍA DE ORDENAMIENTO INTERNO
Métodos: Burbuja, Inserción, Selección, ShellSort, QuickSort, HeapSort, RadixSort
Soporta: 1 archivo (TXT, JSON, Excel .xlsx)
"""

import os
import json
from random import randrange

# =============================================================
# LIBRERÍAS OPCIONALES
# =============================================================

try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

# =============================================================
# MÉTODOS DE ORDENAMIENTO INTERNO
# =============================================================

def burbuja(lista):
    """Ordenamiento por Burbuja - O(n²)"""
    lista = lista[:]
    n = len(lista)
    for i in range(n - 1):
        intercambiado = False
        for j in range(n - 1 - i):
            if lista[j] > lista[j + 1]:
                lista[j], lista[j + 1] = lista[j + 1], lista[j]
                intercambiado = True
        if not intercambiado:
            break
    return lista


def insercion(lista):
    """Ordenamiento por Inserción - O(n²)"""
    lista = lista[:]
    for i in range(1, len(lista)):
        clave = lista[i]
        j = i - 1
        while j >= 0 and lista[j] > clave:
            lista[j + 1] = lista[j]
            j -= 1
        lista[j + 1] = clave
    return lista


def seleccion(lista):
    """Ordenamiento por Selección - O(n²)"""
    lista = lista[:]
    n = len(lista)
    for i in range(n - 1):
        min_idx = i
        for j in range(i + 1, n):
            if lista[j] < lista[min_idx]:
                min_idx = j
        if min_idx != i:
            lista[i], lista[min_idx] = lista[min_idx], lista[i]
    return lista


def shell_sort(lista):
    """ShellSort - O(n log² n) en promedio"""
    lista = lista[:]
    n = len(lista)
    gap = n // 2
    while gap > 0:
        for i in range(gap, n):
            temp = lista[i]
            j = i
            while j >= gap and lista[j - gap] > temp:
                lista[j] = lista[j - gap]
                j -= gap
            lista[j] = temp
        gap //= 2
    return lista


def quick_sort(lista):
    """QuickSort - O(n log n) en promedio"""
    if len(lista) < 2:
        return lista[:]
    pivot_index = randrange(len(lista))
    pivot = lista[pivot_index]
    menores = [x for i, x in enumerate(lista) if x <= pivot and i != pivot_index]
    mayores = [x for i, x in enumerate(lista) if x > pivot]
    return quick_sort(menores) + [pivot] + quick_sort(mayores)


def heapify(lista, n, i):
    """Helper para HeapSort"""
    largest = i
    left = 2 * i + 1
    right = 2 * i + 2
    if left < n and lista[left] > lista[largest]:
        largest = left
    if right < n and lista[right] > lista[largest]:
        largest = right
    if largest != i:
        lista[i], lista[largest] = lista[largest], lista[i]
        heapify(lista, n, largest)


def heap_sort(lista):
    """HeapSort - O(n log n)"""
    lista = lista[:]
    n = len(lista)
    for i in range(n // 2 - 1, -1, -1):
        heapify(lista, n, i)
    for i in range(n - 1, 0, -1):
        lista[0], lista[i] = lista[i], lista[0]
        heapify(lista, i, 0)
    return lista


def radix_sort(lista):
    """RadixSort - O(nk) para enteros no negativos"""
    if not lista:
        return lista[:]
    
    # Verificar si hay números negativos
    if any(x < 0 for x in lista):
        raise ValueError("RadixSort solo funciona con números no negativos")
    
    lista = lista[:]
    max_num = max(lista)
    exp = 1
    RADIX = 10
    
    while max_num // exp > 0:
        buckets = [[] for _ in range(RADIX)]
        for num in lista:
            digit = (num // exp) % RADIX
            buckets[digit].append(num)
        lista = [num for bucket in buckets for num in bucket]
        exp *= RADIX
    
    return lista


# =============================================================
# FUNCIONES DE LECTURA DE ARCHIVOS
# =============================================================

def leer_txt(ruta):
    """Lee números de un archivo TXT"""
    with open(ruta, 'r', encoding='utf-8') as f:
        contenido = f.read()
    
    tokens = contenido.replace(',', ' ').replace('\n', ' ').split()
    numeros = []
    
    for token in tokens:
        try:
            num = float(token)
            numeros.append(int(num) if num == int(num) else num)
        except ValueError:
            print(f"    [!] Advertencia: '{token}' no es número, se omite")
    
    if not numeros:
        raise ValueError("No se encontraron números en el archivo")
    return numeros


def leer_json(ruta):
    """Lee números de un archivo JSON"""
    with open(ruta, 'r', encoding='utf-8') as f:
        datos = json.load(f)
    
    numeros = []
    
    def extraer(obj):
        if isinstance(obj, (int, float)):
            numeros.append(int(obj) if obj == int(obj) else obj)
        elif isinstance(obj, list):
            for item in obj:
                extraer(item)
        elif isinstance(obj, dict):
            for valor in obj.values():
                extraer(valor)
    
    extraer(datos)
    
    if not numeros:
        raise ValueError("No se encontraron números en el JSON")
    return numeros


def leer_excel(ruta):
    """Lee números de un archivo Excel"""
    if not EXCEL_DISPONIBLE:
        raise ImportError("Instala openpyxl: pip install openpyxl")
    
    wb = openpyxl.load_workbook(ruta, data_only=True)
    numeros = []
    
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(values_only=True):
            for celda in fila:
                if isinstance(celda, (int, float)) and celda is not None:
                    num = int(celda) if celda == int(celda) else celda
                    numeros.append(num)
    wb.close()
    
    if not numeros:
        raise ValueError("No se encontraron números en el Excel")
    return numeros


def leer_archivo(ruta):
    """Detecta el tipo de archivo y lee los números"""
    ext = os.path.splitext(ruta)[1].lower()
    
    if ext == '.txt':
        return leer_txt(ruta)
    elif ext == '.json':
        return leer_json(ruta)
    elif ext in ('.xlsx', '.xls'):
        return leer_excel(ruta)
    else:
        raise ValueError(f"Formato no soportado: {ext}. Usa .txt, .json o .xlsx")


# =============================================================
# DICCIONARIO DE MÉTODOS
# =============================================================

METODOS_INTERNOS = {
    "burbuja": burbuja,
    "insercion": insercion,
    "seleccion": seleccion,
    "shellsort": shell_sort,
    "quicksort": quick_sort,
    "heapsort": heap_sort,
    "radixsort": radix_sort
}

NOMBRES_METODOS_INTERNOS = {
    "burbuja": "Burbuja",
    "insercion": "Inserción",
    "seleccion": "Selección",
    "shellsort": "ShellSort",
    "quicksort": "QuickSort",
    "heapsort": "HeapSort",
    "radixsort": "RadixSort"
}
