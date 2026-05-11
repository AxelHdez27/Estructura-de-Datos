"""
PROGRAMA PRINCIPAL - EJERCICIOS DE ORDENAMIENTO
Integra TODOS los métodos: Burbuja, Inserción, Selección, Radix, Heap, Quick, Shell
Y métodos externos: Intercalación, Mezcla Directa, Mezcla Equilibrada
"""

import os
import json
import time
import random
from random import randrange, randint

# =============================================================
# LIBRERÍAS OPCIONALES
# =============================================================

try:
    import openpyxl
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

try:
    import tkinter as tk
    from tkinter import filedialog
    TKINTER_DISPONIBLE = True
except ImportError:
    TKINTER_DISPONIBLE = False


# =============================================================
# MÉTODOS DE ORDENAMIENTO INTERNO (COMPLETOS)
# =============================================================

# 1. Burbuja
def bubble_sort(arr):
    arr = arr[:]
    n = len(arr)
    for i in range(n - 1):
        swapped = False
        for j in range(n - 1 - i):
            if arr[j] > arr[j + 1]:
                arr[j], arr[j + 1] = arr[j + 1], arr[j]
                swapped = True
        if not swapped:
            break
    return arr


# 2. Inserción
def insertion_sort(arr):
    arr = arr[:]
    for i in range(1, len(arr)):
        key = arr[i]
        j = i - 1
        while j >= 0 and arr[j] > key:
            arr[j + 1] = arr[j]
            j -= 1
        arr[j + 1] = key
    return arr


# 3. Selección
def selection_sort(arr):
    arr = arr[:]
    n = len(arr)
    for i in range(n - 1):
        min_idx = i
        for j in range(i + 1, n):
            if arr[j] < arr[min_idx]:
                min_idx = j
        if min_idx != i:
            arr[i], arr[min_idx] = arr[min_idx], arr[i]
    return arr


# 4. Radix Sort
RADIX = 10
def radix_sort(arr):
    if not arr:
        return []
    # Convertir a enteros positivos si es necesario
    arr = [int(x) for x in arr]
    arr = arr[:]
    placement = 1
    max_digit = max(arr)
    while placement <= max_digit:
        buckets = [[] for _ in range(RADIX)]
        for i in arr:
            tmp = int((i / placement) % RADIX)
            buckets[tmp].append(i)
        a = 0
        for b in range(RADIX):
            for i in buckets[b]:
                arr[a] = i
                a += 1
        placement *= RADIX
    return arr


# 5. Heap Sort
def heapify(arr, index, heap_size):
    largest = index
    left = 2 * index + 1
    right = 2 * index + 2
    if left < heap_size and arr[left] > arr[largest]:
        largest = left
    if right < heap_size and arr[right] > arr[largest]:
        largest = right
    if largest != index:
        arr[largest], arr[index] = arr[index], arr[largest]
        heapify(arr, largest, heap_size)


def heap_sort(arr):
    arr = arr[:]
    n = len(arr)
    for i in range(n // 2 - 1, -1, -1):
        heapify(arr, i, n)
    for i in range(n - 1, 0, -1):
        arr[0], arr[i] = arr[i], arr[0]
        heapify(arr, 0, i)
    return arr


# 6. Quick Sort
def quick_sort(arr):
    if len(arr) < 2:
        return arr[:]
    arr = arr[:]
    pivot_index = randrange(len(arr))
    pivot = arr.pop(pivot_index)
    lesser = [item for item in arr if item <= pivot]
    greater = [item for item in arr if item > pivot]
    return [*quick_sort(lesser), pivot, *quick_sort(greater)]


# 7. Shell Sort
def shell_sort(arr):
    arr = arr[:]
    gaps = [701, 301, 132, 57, 23, 10, 4, 1]
    for gap in gaps:
        for i in range(gap, len(arr)):
            insert_value = arr[i]
            j = i
            while j >= gap and arr[j - gap] > insert_value:
                arr[j] = arr[j - gap]
                j -= gap
            if j != i:
                arr[j] = insert_value
    return arr


# =============================================================
# MÉTODOS DE ORDENAMIENTO EXTERNO
# =============================================================

def intercalar(izq, der):
    resultado = []
    i = j = 0
    while i < len(izq) and j < len(der):
        if izq[i] <= der[j]:
            resultado.append(izq[i])
            i += 1
        else:
            resultado.append(der[j])
            j += 1
    resultado.extend(izq[i:])
    resultado.extend(der[j:])
    return resultado


def ordenar_intercalacion(lista):
    if len(lista) <= 1:
        return lista[:]
    medio = len(lista) // 2
    return intercalar(
        ordenar_intercalacion(lista[:medio]),
        ordenar_intercalacion(lista[medio:])
    )


def mezclar_runs(lista, inicio, medio, fin):
    izq = lista[inicio:medio]
    der = lista[medio:fin]
    i = j = 0
    k = inicio
    while i < len(izq) and j < len(der):
        if izq[i] <= der[j]:
            lista[k] = izq[i]
            i += 1
        else:
            lista[k] = der[j]
            j += 1
        k += 1
    while i < len(izq):
        lista[k] = izq[i]
        i += 1
        k += 1
    while j < len(der):
        lista[k] = der[j]
        j += 1
        k += 1


def ordenar_mezcla_directa(lista):
    lista = lista[:]
    n = len(lista)
    tam = 1
    while tam < n:
        for inicio in range(0, n, tam * 2):
            medio = min(inicio + tam, n)
            fin = min(inicio + tam * 2, n)
            if medio < fin:
                mezclar_runs(lista, inicio, medio, fin)
        tam *= 2
    return lista


def ordenar_mezcla_equilibrada(lista, k=2):
    if len(lista) <= 1:
        return lista[:]
    runs = [[x] for x in lista]
    while len(runs) > 1:
        nuevos_runs = []
        for i in range(0, len(runs), k):
            grupo = runs[i:i + k]
            mezclado = grupo[0]
            for r in grupo[1:]:
                mezclado = intercalar(mezclado, r)
            nuevos_runs.append(mezclado)
        runs = nuevos_runs
    return runs[0]


# =============================================================
# LECTURA DE ARCHIVOS
# =============================================================

def leer_txt(ruta):
    with open(ruta, "r", encoding="utf-8") as f:
        contenido = f.read()
    tokens = contenido.replace(",", " ").replace("\n", " ").split()
    numeros = []
    for tok in tokens:
        tok = tok.strip()
        if not tok:
            continue
        try:
            num = float(tok)
            numeros.append(int(num) if num == int(num) else num)
        except ValueError:
            raise ValueError(f"Valor no numérico: '{tok}'")
    return numeros


def leer_json(ruta):
    with open(ruta, "r", encoding="utf-8") as f:
        datos = json.load(f)

    def extraer_numeros(obj):
        encontrados = []
        if isinstance(obj, list):
            for item in obj:
                if isinstance(item, (int, float)):
                    encontrados.append(int(item) if item == int(item) else item)
                elif isinstance(item, (list, dict)):
                    encontrados.extend(extraer_numeros(item))
        elif isinstance(obj, dict):
            for val in obj.values():
                encontrados.extend(extraer_numeros(val))
        return encontrados

    numeros = extraer_numeros(datos)
    if not numeros:
        raise ValueError(f"No se encontraron números en '{ruta}'.")
    return numeros


def leer_excel(ruta):
    if not EXCEL_DISPONIBLE:
        raise ImportError("Instala openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(ruta, data_only=True)
    numeros = []
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(values_only=True):
            for celda in fila:
                if isinstance(celda, (int, float)) and celda is not None:
                    num = int(celda) if celda == int(celda) else celda
                    numeros.append(num)
    wb.close()
    if not numeros:
        raise ValueError(f"No se encontraron números en '{ruta}'.")
    return numeros


def leer_archivo(ruta):
    ext = os.path.splitext(ruta)[1].lower()
    if ext == ".txt":
        return leer_txt(ruta)
    elif ext in (".xlsx", ".xls"):
        return leer_excel(ruta)
    elif ext == ".json":
        return leer_json(ruta)
    else:
        raise ValueError(f"Formato no soportado: '{ext}'")


def pedir_ruta(numero):
    while True:
        print(f"\n  Archivo {numero} — ¿Cómo quieres buscarlo?")
        print("    E = Abrir explorador de archivos")
        print("    M = Escribir la ruta manualmente")
        modo = input("  Elige (E/M): ").strip().upper()

        if modo == "E":
            if not TKINTER_DISPONIBLE:
                print("  [!] tkinter no disponible. Usa modo manual (M).")
                continue
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            ruta = filedialog.askopenfilename(
                title=f"Selecciona el Archivo {numero}",
                filetypes=[("Todos", "*.txt *.xlsx *.xls *.json"), 
                          ("Texto", "*.txt"), ("Excel", "*.xlsx"), ("JSON", "*.json")]
            )
            root.destroy()
            if not ruta:
                print("  [!] No seleccionaste ningún archivo.")
                continue
            return ruta
        elif modo == "M":
            ruta = input(f"  Ruta del archivo {numero}: ").strip().strip('"')
            if not ruta:
                continue
            if not os.path.isfile(ruta):
                print(f"  [!] Archivo no encontrado")
                continue
            return ruta
        else:
            print("  [!] Opción inválida.")


def pedir_ruta_salida():
    while True:
        print("\n  Archivo de salida — ¿Cómo quieres guardarlo?")
        print("    E = Abrir explorador para guardar")
        print("    M = Escribir nombre manualmente")
        modo = input("  Elige (E/M): ").strip().upper()

        if modo == "E":
            if not TKINTER_DISPONIBLE:
                print("  [!] tkinter no disponible. Usa modo manual (M).")
                continue
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            ruta = filedialog.asksaveasfilename(
                title="Guardar archivo ordenado",
                defaultextension=".txt",
                filetypes=[("Archivo de texto", "*.txt")]
            )
            root.destroy()
            if not ruta:
                continue
            return ruta
        elif modo == "M":
            ruta = input("  Nombre del archivo (ej: resultado.txt): ").strip()
            if not ruta.endswith(".txt"):
                ruta += ".txt"
            return ruta
        else:
            print("  [!] Opción inválida.")


def crear_archivos_ejemplo():
    print("\n  Creando archivos de ejemplo...")
    with open("ejemplo1.txt", "w") as f:
        f.write("85, 72, 91, 60, 78, 34, 55, 19")
    with open("ejemplo2.txt", "w") as f:
        f.write("43, 67, 11, 98, 25, 50, 73, 88")
    with open("ejemplo1.json", "w") as f:
        json.dump([85, 72, 91, 60, 78, 34, 55, 19], f)
    with open("ejemplo2.json", "w") as f:
        json.dump({"numeros": [43, 67, 11, 98, 25, 50, 73, 88]}, f)
    if EXCEL_DISPONIBLE:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Numeros"
        for v in [85, 72, 91, 60, 78, 34, 55, 19]:
            ws.append([v])
        wb.save("ejemplo1.xlsx")
        print("  ✅ Creados: ejemplo1.txt, ejemplo2.txt, ejemplo1.json, ejemplo2.json, ejemplo1.xlsx")
    else:
        print("  ✅ Creados: ejemplo1.txt, ejemplo2.txt, ejemplo1.json, ejemplo2.json")


# =============================================================
# EJECUTAR MÉTODOS
# =============================================================

def ejecutar_metodo(nombre_metodo, funcion_ordenamiento, es_externo=False, k=None):
    print(f"\n{'─'*60}")
    print(f"  MÉTODO: {nombre_metodo}")
    print(f"{'─'*60}")
    
    ruta1 = pedir_ruta(1)
    ruta2 = pedir_ruta(2)
    
    try:
        nums1 = leer_archivo(ruta1)
        nums2 = leer_archivo(ruta2)
    except Exception as e:
        print(f"\n  [ERROR] {e}")
        return
    
    combinados = nums1 + nums2
    
    print(f"\n  📁 {os.path.basename(ruta1)} → {nums1[:10]}{'...' if len(nums1)>10 else ''}")
    print(f"  📁 {os.path.basename(ruta2)} → {nums2[:10]}{'...' if len(nums2)>10 else ''}")
    print(f"  📊 Total: {len(combinados)} números")
    
    if es_externo and k:
        resultado = funcion_ordenamiento(combinados, k)
    else:
        resultado = funcion_ordenamiento(combinados)
    
    print(f"  ✅ Resultado: {resultado[:20]}{'...' if len(resultado)>20 else ''}")
    
    ruta_salida = pedir_ruta_salida()
    try:
        with open(ruta_salida, "w", encoding="utf-8") as f:
            f.write(f"Método: {nombre_metodo}\n")
            f.write(f"Archivo 1: {os.path.basename(ruta1)}\n")
            f.write(f"Archivo 2: {os.path.basename(ruta2)}\n")
            f.write(f"Total números: {len(resultado)}\n")
            f.write("="*50 + "\n")
            f.write(", ".join(str(n) for n in resultado))
        print(f"\n  💾 Guardado en: {ruta_salida}")
    except Exception as e:
        print(f"\n  [ERROR] No se pudo guardar: {e}")


def pedir_k():
    while True:
        k_str = input("  Número de cintas k [default=2]: ").strip()
        if k_str == "":
            return 2
        try:
            k = int(k_str)
            if k >= 2:
                return k
            print("  [!] k debe ser >= 2.")
        except ValueError:
            print("  [!] Ingresa un número entero.")


def limpiar():
    os.system('cls' if os.name == 'nt' else 'clear')


def mostrar_banner():
    explorador = "✅ SI" if TKINTER_DISPONIBLE else "❌ NO"
    excel = "✅ SI" if EXCEL_DISPONIBLE else "❌ NO"
    print("=" * 60)
    print("   ORDENAMIENTO EXTERNO — TXT / EXCEL / JSON")
    print("=" * 60)
    print(f"   Explorador visual : {explorador}")
    print(f"   Soporte Excel     : {excel}")
    print("-" * 60)
    print("\n   📌 EJERCICIOS REQUERIDOS (Ordenamiento Interno):")
    print("     1. Método Burbuja")
    print("     2. Método Inserción")
    print("     3. Método Selección")
    print("     4. Radix Sort ✨ NUEVO")
    print("     5. Heap Sort ✨ NUEVO")
    print("     6. Quick Sort ✨ NUEVO")
    print("     7. Shell Sort ✨ NUEVO")
    print("\n   📌 MÉTODOS EXTERNOS:")
    print("     8. Intercalación (Merge Sort)")
    print("     9. Mezcla Directa")
    print("    10. Mezcla Equilibrada (k cintas)")
    print("\n   📌 UTILIDADES:")
    print("    11. Crear archivos de ejemplo")
    print("     0. Salir")
    print("-" * 60)


# =============================================================
# MENÚ PRINCIPAL
# =============================================================

def main():
    limpiar()
    print("=" * 60)
    print("   SISTEMA DE ORDENAMIENTO - EJERCICIOS INTEGRADOS")
    print("=" * 60)
    print("  Lee dos archivos, mezcla sus números con el")
    print("  método elegido y genera un nuevo .txt ordenado.\n")
    
    respuesta = input("  ¿Crear archivos de ejemplo? (s/n): ").strip().lower()
    if respuesta == "s":
        crear_archivos_ejemplo()
    
    input("\n  Presiona ENTER para continuar al menú...")
    
    while True:
        limpiar()
        mostrar_banner()
        opcion = input("  Selecciona una opción: ").strip()
        
        if opcion == "0":
            print("\n  👋 ¡Hasta luego!\n")
            break
        elif opcion == "1":
            ejecutar_metodo("Burbuja", bubble_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "2":
            ejecutar_metodo("Inserción", insertion_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "3":
            ejecutar_metodo("Selección", selection_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "4":
            ejecutar_metodo("Radix Sort", radix_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "5":
            ejecutar_metodo("Heap Sort", heap_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "6":
            ejecutar_metodo("Quick Sort", quick_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "7":
            ejecutar_metodo("Shell Sort", shell_sort)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "8":
            ejecutar_metodo("Intercalación", ordenar_intercalacion)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "9":
            ejecutar_metodo("Mezcla Directa", ordenar_mezcla_directa)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "10":
            k = pedir_k()
            ejecutar_metodo(f"Mezcla Equilibrada (k={k})", ordenar_mezcla_equilibrada, es_externo=True, k=k)
            input("\n  Presiona ENTER para volver al menú...")
        elif opcion == "11":
            crear_archivos_ejemplo()
            input("\n  Presiona ENTER para volver al menú...")
        else:
            print("\n  [!] Opción no válida. Elige 0-11.")
            input("\n  Presiona ENTER para continuar...")


if __name__ == "__main__":
    main()
